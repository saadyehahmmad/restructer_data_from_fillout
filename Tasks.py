import os
from flask import Flask, request, redirect, url_for, send_file, render_template
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Alignment, Font

app = Flask(__name__)

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Ensure the folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Step 1: Delete the last two columns
    max_col = ws.max_column
    ws.delete_cols(max_col)
    ws.delete_cols(max_col - 1)

    # Step 2: Delete the first two columns
    ws.delete_cols(1)
    ws.delete_cols(1)

    # Step 3: Insert two columns for counting 1's and 0's
    max_row = ws.max_row
    ws.insert_cols(1, 2)
    ws.cell(row=1, column=1, value="Count of 1's")
    ws.cell(row=1, column=2, value="Count of 0's")

    for row in range(2, max_row + 1):
        ones_count = sum([1 for cell in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=row, max_row=row) if cell[0].value == 1])
        zeros_count = sum([1 for cell in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=row, max_row=row) if cell[0].value == 0])
        ws.cell(row=row, column=1, value=ones_count)
        ws.cell(row=row, column=2, value=zeros_count)

    # Step 4: Delete all columns that don't have a value in the second row or below
    max_col = ws.max_column
    for col in range(max_col, 0, -1):
        if all([ws.cell(row=row, column=col).value is None for row in range(2, max_row + 1)]):
            ws.delete_cols(col)

    # Step 5: Transpose the table
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column, values_only=True):
        data.append(row)

    transposed_data = list(zip(*data))  # Transpose the data using zip

    # Clear the original data
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Write the transposed data starting from row 1
    for r_idx, row in enumerate(transposed_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Update max_row and max_col after transpose
    max_row = ws.max_row
    max_col = ws.max_column

    # Step 6: Hide the rows where all columns contain 1's or are empty
    for row in range(2, max_row + 1):  # Start from the second row
        all_ones = True
        for col in range(2, max_col + 1):  # Check each column in the row starting from the second column
            cell_value = ws.cell(row=row, column=col).value
            if cell_value != 1 and cell_value is not None:  # If the cell is neither 1 nor empty, set all_ones to False
                all_ones = False
                break
        if all_ones:
            ws.row_dimensions[row].hidden = True  # Hide the row if all values are 1 or empty

    # Step 7: Set right-to-left text direction and align right
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', readingOrder=2)  # Right align and set text direction to RTL

    # Step 8: Make the first column bold with font size 14
    for row in range(1, max_row + 1):  # Iterate over all rows
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, size=14)

    # Step 9: Autofit column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Step 10: Make the document right to left
    ws.sheet_view.rightToLeft = True

    # Step 11: Save the processed workbook
    processed_file_path = os.path.join(app.config['PROCESSED_FOLDER'], os.path.basename(file_path))
    wb.save(processed_file_path)
    return processed_file_path


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(file_path)
            try:
                restructured_file_path = process_excel(file_path)
                return send_file(restructured_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return render_template('upload.html')


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5000)
