from flask import Flask, request, send_file, render_template, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Configure upload folders
UPLOAD_FOLDER_1 = 'KPI'
UPLOAD_FOLDER_2 = 'Tasks'
app.config['UPLOAD_FOLDER_1'] = UPLOAD_FOLDER_1
app.config['UPLOAD_FOLDER_2'] = UPLOAD_FOLDER_2
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

# Ensure the folders exist
os.makedirs(UPLOAD_FOLDER_1, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_2, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Flask Application 1
@app.route('/kpi', methods=['GET', 'POST'])
def app1():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER_1'], secure_filename(file.filename))
            file.save(file_path)
            try:
                restructured_file_path = restructure_excel(file_path)  # This function is from your first app
                return send_file(restructured_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return render_template('upload1.html')  # Template for the first application

def restructure_excel(file_path):
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls)

    # Define roles
    roles = ['الهيد', 'الكاشير', 'باريستا(1)', 'باريستا(2)', 'باريستا(3)', 'باريستا(4)', 'صالة(1)', 'صالة(2)', 'صالة(3)']

    kpi_criteria = [
        "1- الالتزام بمواعيد الحضور والغياب",
        "2- الالتزام باليونيفورم واللباس المطلوب من حيث نظافته ، العطر , الأظافر، الشعر ، النظافة الشخصية ، والاكسسوارات",
        "3- الالتزام بمواعيد البريكات وترك غرفة البريك نظيفة",
        "4- عدم استخدام الهاتف الشخصي وان يكون هاتفه متاح بعد الدوام",
        "5- هل يحترم زملائه ويجيد التعامل معهم ولا يتسبب في حدوث مشاكل بين زملائه",
        "6- هل يحتاج إلى التوجيه والرقابة المستمرة أو قادر على إنجاز أعماله دون انتظار التعليمات والتوجيهات",
        "7- هل تطورت مهاراته الوظيفية من حيث السرعة والإتقان",
        "8- هل يتأقلم ويتقبل التعديلات والتغييرات بسرعة؟ إلى أي مدى يتقبل ويستوعب تعليقات وملاحظات مديريه",
        "9- رفع التقارير وإخبار مديره بالمعلومات ونقل المشاكل والفيدباك بشكل مستمر",
        "10- الالتزام بمهام التشطيب وجودة أداء المهام المطلوبة",
        "11- التباهي بالسلوكيات الخاطئة (سهرات او انشطة مشبوهة بعد الدوام)",
        "12- مدى التركيز والانتباه على الصالة وراحة العملاء وأي مشاكل أو خلل وسرعة التعامل معها وإيصالها و التركيز على حلها وعدم نسيانها",
        "13- مدى الالتزام بالألقاب مع المسؤول (سير)",
        "14- سرعة أخذ الطلبات وإنجازها 2-3 دقائق والتأكد من نظافة الطلب ومطابقة المعايير من حيث نظافة الكاسة والالتزام بالريسيبي",
        "15- تشطيب البار وتنظيفه خلال 45 دقيقة",
        "16- استقبال الضيوف وتشمل الموظفين القدامى",
        "17- حفظ أسماء العملاء وألقابهم مع الحفاظ على طبيعة العلاقة friendly ليس friends والالتزام والحفاظ على العلاقات الرسمية معهم ومدة الوقوف معهم وتجنب الأحاديث الجانبية",
        "18- الحفاظ على الوقفة السليمة والترحيب بالعملاء",
        "19- مراقبة الطلبات المرتجعة أو لم يتم شربها وعمل تقرير بها ومعرفة المشكلة ضبط جودة المشروبات والطلبات عند التحضير وقبل التسليم",
        "20- القدرة على حل المشاكل سواء مشاكل الموظفين الخاصة  و ومشاكل بين الموظفين خلال العمل و توزيع البريكات على الموظفين بشكل عادل",
        "21- مدى الاتكالية في المهام وسوء استخدام السلطة مع الموظفين",
        "22- فهم الواجبات والمهام المطلوبة منه بشكل واضح وكيفية تطبيقها و القدرة على تدريب موظفيه ورفع أدائهم ومعرفتهم في الباريستا",
        "23- الصيانات داخل البار ومتابعتها حتى يتم إنجازها",
        "24- القدرة على طلب الطلبيات والحفاظ على الـstock المناسب ومتابعة المخزون وقت وصول الطلبية والإبلاغ في حال كان هناك نقص أو مشكلة أو تأخير في وقت الوصول",
        "25- قدرة الموظفين على العمل دون إشراف ورقابة وتوجيه دائم ضبط سلوكهم وتحركاتهم داخل البار",
        "END"
    ]


    def get_columns_for_role(role):
        return [col for col in df.columns if role in col]

    additional_columns = ['Branch', 'اسم المسؤول السوبرفايزر', 'الشفت', 'Submission started']

    # Ensure the additional columns exist in the DataFrame
    additional_columns = [col for col in additional_columns if col in df.columns]

    def create_role_data(index, max_length, role_columns):
        role_data = [None] * max_length
        for col in role_columns:
            try:
                col_index = int(col.split('-')[0].strip())
                if col_index <= max_length:
                    role_data[col_index - 1] = df.at[index, col]
            except ValueError:
                continue
        return role_data

    def sanitize_sheet_name(name):
        return str(name).replace(':', '-').replace('/', '-').replace('\\', '-').replace('*', '-').replace('?',
                                                                                                          '-').replace(
            '[', '-').replace(']', '-')

    output_path_custom = 'kpi.xlsx'

    with pd.ExcelWriter(output_path_custom, engine='openpyxl') as writer:
        sheet_created = False
        for index, row in df.iterrows():
            try:
                sheet_name = sanitize_sheet_name(row['اسم المسؤول السوبرفايزر'])

                max_length = len(kpi_criteria)
                data = {'KPI Criteria': kpi_criteria}

                for role in roles:
                    role_columns = get_columns_for_role(role)
                    role_data = create_role_data(index, max_length, role_columns)
                    data[role] = role_data

                restructured_df = pd.DataFrame(data)

                averages = (restructured_df.apply(pd.to_numeric, errors='coerce').mean()) / 0.03
                averages_row = pd.DataFrame(averages).T

                if additional_columns:
                    additional_data = row[additional_columns].values
                    additional_df = pd.DataFrame([additional_data], columns=additional_columns)
                    additional_df.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)

                dynamic_header = ['']
                for role in roles:
                    role_name_col = ""
                    for col in row.keys():
                        if isinstance(col, str) and "اسم" in col and role in col:
                            role_name_col = col
                            break
                    dynamic_header.append(row[role_name_col] if role_name_col else "")
                header_df = pd.DataFrame([dynamic_header])
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False, header=False)

                restructured_df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)

                averages_row.to_excel(writer, sheet_name=sheet_name, startrow=len(restructured_df) + 3, header=False,
                                      index=False)

                worksheet = writer.sheets[sheet_name]

                # Set modern font and alignment
                modern_font = Font(name='Calibri', size=12, bold=True)
                header_font = Font(name='Calibri', size=12, bold=True)
                alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))

                # Color header
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                for col_index, cell in enumerate(worksheet[2], start=1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border

                # Color additional columns
                for col_index, col_name in enumerate(additional_columns, start=1):
                    cell = worksheet.cell(row=1, column=col_index)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border
                    value = row[col_name]
                    value_cell = worksheet.cell(row=2, column=col_index)
                    value_cell.value = value
                    value_cell.font = modern_font
                    value_cell.alignment = alignment
                    value_cell.border = border

                # Format all other cells
                for row in worksheet.iter_rows(min_row=3, max_row=len(restructured_df) + 3, min_col=1,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.font = modern_font
                        cell.alignment = alignment
                        cell.border = border

                # Align Index column to the left
                for cell in worksheet['A']:
                    cell.alignment = left_alignment

                highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for cell in worksheet.iter_rows(min_row=len(restructured_df) + 4, max_row=len(restructured_df) + 4,
                                                min_col=1, max_col=worksheet.max_column):
                    for c in cell:
                        c.fill = highlight_fill
                        c.font = modern_font
                        c.alignment = alignment
                        c.border = border

                # Autofit columns
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

                sheet_created = True
            except Exception as e:
                print(f"An error occurred while processing row {index}: {e}")
                continue

        if not sheet_created:
            # Create an empty sheet to avoid IndexError
            pd.DataFrame().to_excel(writer, sheet_name='EmptySheet')

    return output_path_custom


# Flask Application 2
@app.route('/tasks', methods=['GET', 'POST'])
def app2():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER_2'], secure_filename(file.filename))
            file.save(file_path)
            try:
                processed_file_path = process_excel(file_path)  # This function is from your second app
                return send_file(processed_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return render_template('upload2.html')  # Template for the second application

def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Step 1: Delete the last two columns
    max_col = ws.max_column
    ws.delete_cols(max_col)
    ws.delete_cols(max_col - 1)

    # Step 2: Delete the first two columns
    ws.delete_cols(1)
    ws.delete_cols(1)

    # Step 3: Insert two columns for counting 1's and 0's
    max_row = ws.max_row
    ws.insert_cols(1, 2)
    ws.cell(row=1, column=1, value="Count of 1's")
    ws.cell(row=1, column=2, value="Count of 0's")

    for row in range(2, max_row + 1):
        ones_count = sum([1 for cell in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=row, max_row=row) if cell[0].value == 1])
        zeros_count = sum([1 for cell in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=row, max_row=row) if cell[0].value == 0])
        ws.cell(row=row, column=1, value=ones_count)
        ws.cell(row=row, column=2, value=zeros_count)

    # Step 4: Delete all columns that don't have a value in the second row or below
    max_col = ws.max_column
    for col in range(max_col, 0, -1):
        if all([ws.cell(row=row, column=col).value is None for row in range(2, max_row + 1)]):
            ws.delete_cols(col)

    # Step 5: Transpose the table
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column, values_only=True):
        data.append(row)

    transposed_data = list(zip(*data))  # Transpose the data using zip

    # Clear the original data
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Write the transposed data starting from row 1
    for r_idx, row in enumerate(transposed_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Update max_row and max_col after transpose
    max_row = ws.max_row
    max_col = ws.max_column

    # Step 6: Hide the rows where all columns contain 1's or are empty
    for row in range(2, max_row + 1):  # Start from the second row
        all_ones = True
        for col in range(2, max_col + 1):  # Check each column in the row starting from the second column
            cell_value = ws.cell(row=row, column=col).value
            if cell_value != 1 and cell_value is not None:  # If the cell is neither 1 nor empty, set all_ones to False
                all_ones = False
                break
        if all_ones:
            ws.row_dimensions[row].hidden = True  # Hide the row if all values are 1 or empty

    # Step 7: Set right-to-left text direction and align right
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', readingOrder=2)  # Right align and set text direction to RTL

    # Step 8: Make the first column bold with font size 14
    for row in range(1, max_row + 1):  # Iterate over all rows
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, size=14)

    # Step 9: Autofit column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Step 10: Make the document right to left
    ws.sheet_view.rightToLeft = True

    # Step 11: Save the processed workbook
    processed_file_path = os.path.join(app.config['PROCESSED_FOLDER'], os.path.basename(file_path))
    wb.save(processed_file_path)
    return processed_file_path

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
