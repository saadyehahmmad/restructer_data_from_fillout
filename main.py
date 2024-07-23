from flask import Flask, request, send_file, render_template_string
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def restructure_excel(file_path):
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls)

    head_barista_columns = [col for col in df.columns if 'هيد' in col]
    cashier_columns = [col for col in df.columns if 'كاشير' in col]
    barista_1_columns = [col for col in df.columns if 'باريستا(1)' in col]
    barista_2_columns = [col for col in df.columns if 'باريستا(2)' in col]
    barista_3_columns = [col for col in df.columns if 'باريستا(3)' in col]
    barista_4_columns = [col for col in df.columns if 'باريستا(4)' in col]
    host_1_columns = [col for col in df.columns if 'صالة(1)' in col]
    host_2_columns = [col for col in df.columns if 'صالة(2)' in col]
    host_3_columns = [col for col in df.columns if 'صالة(3)' in col]

    additional_columns = ['Branch', 'اسم المسؤول السوبرفايزر', 'الشفت', 'Submission started']

    # Ensure the additional columns exist in the DataFrame
    additional_columns = [col for col in additional_columns if col in df.columns]

    def pad_array(arr, length):
        return np.pad(arr, (0, length - len(arr)), constant_values=None)

    def sanitize_sheet_name(name):
        return str(name).replace(':', '-').replace('/', '-').replace('\\', '-').replace('*', '-').replace('?',
                                                                                                          '-').replace(
            '[', '-').replace(']', '-')

    output_path_custom = 'restructured_KPI_custom.xlsx'

    with pd.ExcelWriter(output_path_custom, engine='openpyxl') as writer:
        sheet_created = False
        for index, row in df.iterrows():
            sheet_name = sanitize_sheet_name(row['اسم المسؤول السوبرفايزر'])

            head_barista_data = row[head_barista_columns].values.astype(object)
            cashier_data = row[cashier_columns].values.astype(object)
            barista_1_data = row[barista_1_columns].values.astype(object)
            barista_2_data = row[barista_2_columns].values.astype(object)
            barista_3_data = row[barista_3_columns].values.astype(object)
            barista_4_data = row[barista_4_columns].values.astype(object)
            host_1_data = row[host_1_columns].values.astype(object)
            host_2_data = row[host_2_columns].values.astype(object)
            host_3_data = row[host_3_columns].values.astype(object)

            max_length = max(len(head_barista_data), len(cashier_data), len(barista_1_data), len(barista_2_data),
                             len(barista_3_data), len(barista_4_data), len(host_1_data), len(host_2_data),
                             len(host_3_data))

            head_barista_data = pad_array(head_barista_data, max_length)
            cashier_data = pad_array(cashier_data, max_length)
            barista_1_data = pad_array(barista_1_data, max_length)
            barista_2_data = pad_array(barista_2_data, max_length)
            barista_3_data = pad_array(barista_3_data, max_length)
            barista_4_data = pad_array(barista_4_data, max_length)
            host_1_data = pad_array(host_1_data, max_length)
            host_2_data = pad_array(host_2_data, max_length)
            host_3_data = pad_array(host_3_data, max_length)

            data = {
                'Head Barista': head_barista_data,
                'Cashier': cashier_data,
                'Barista (1)': barista_1_data,
                'Barista (2)': barista_2_data,
                'Barista (3)': barista_3_data,
                'Barista (4)': barista_4_data,
                'Host (1)': host_1_data,
                'Host (2)': host_2_data,
                'Host (3)': host_3_data
            }
            restructured_df = pd.DataFrame(data)

            averages = restructured_df.apply(pd.to_numeric, errors='coerce').mean()
            averages_row = pd.DataFrame(averages).T

            if additional_columns:
                additional_data = row[additional_columns].values
                additional_df = pd.DataFrame([additional_data], columns=additional_columns)
                additional_df.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)

            restructured_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

            averages_row.to_excel(writer, sheet_name=sheet_name, startrow=len(restructured_df) + 2, header=False,
                                  index=False)

            worksheet = writer.sheets[sheet_name]

            highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for cell in worksheet.iter_rows(min_row=len(restructured_df) + 3, max_row=len(restructured_df) + 3,
                                            min_col=1, max_col=worksheet.max_column):
                for c in cell:
                    c.fill = highlight_fill

            for col_index, col_name in enumerate(additional_columns, start=1):
                cell = worksheet.cell(row=1, column=col_index)
                cell.font = Font(size=14, bold=True)
                value = row[col_name]
                value_cell = worksheet.cell(row=2, column=col_index)
                value_cell.value = value
                value_cell.font = Font(size=14)

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            sheet_created = True

        if not sheet_created:
            # Create an empty sheet to avoid IndexError
            pd.DataFrame().to_excel(writer, sheet_name='EmptySheet')

    return output_path_custom


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(file_path)
            try:
                restructured_file_path = restructure_excel(file_path)
                return send_file(restructured_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return '''
<!doctype html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>رفع ملف جديد</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f5f5f5;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin: 0;
        }
        .container {
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            text-align: center;
        }
        h1, h2 {
            margin: 0 0 20px 0;
        }
        h2 strong {
            color: red;
        }
        form {
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        input[type="file"] {
            background-color: tomato;
            color: white;
            border: none;
            padding:5px 20px 5px 20px;
            margin:50px;
            width:400px;
            font-size: 20px;
            border-radius: 4px;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
         input[type="file"]:hover {
            background-color: black;

        }
        button[type="submit"] {
            background-color: #007bff;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            height:50px;
            padding:5px 190px 5px 190px;
            font-size:14px;
            transition: background-color 0.3s ease;
        }
        button[type="submit"]:hover {
            background-color: #0056b3;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>رفع ملف جديد</h1>
        <h2>تأكد من أن صيغة الجدول هي <strong>.xlsx</strong> قبل الرفع.</h2>
        <form method="post" enctype="multipart/form-data">
           <p>ارفق الملف لتحويله</p> <input type="file" name="file" accept=".xlsx" >
            <button type="submit">تحميل</button>
        </form>
    </div>
</body>
</html>

    '''


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5000, debug=True)
