from flask import Flask, request, send_file, render_template_string
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def restructure_excel(file_path):
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls)

    # Define roles
    roles = ['هيد', 'كاشير', 'باريستا(1)', 'باريستا(2)', 'باريستا(3)', 'باريستا(4)', 'صالة(1)', 'صالة(2)', 'صالة(3)']

    kpi_criteria = [
        "1- الالتزام بمواعيد الحضور والغياب",
        "2- الالتزام باليونيفورم واللباس المطلوب من حيث نظافته ، العطر , الأظافر، الشعر ، النظافة الشخصية ، والاكسسوارات",
        "3- الالتزام بمواعيد البريكات وترك غرفة البريك نظيفة",
        "4- عدم استخدام الهاتف الشخصي وان يكون هاتفه متاح بعد الدوام",
        "5- هل يحترم زملائه ويجيد التعامل معهم ولا يتسبب في حدوث مشاكل بين زملائه",
        "6- هل يحتاج إلى التوجيه والرقابة المستمرة أو قادر على إنجاز أعماله دون انتظار التعليمات والتوجيهات",
        "7- هل تطورت مهاراته الوظيفية من حيث السرعة والإتقان",
        "8- هل يتأقلم ويتقبل التعديلات والتغييرات بسرعة؟ إلى أي مدى يتقبل ويستوعب تعليقات وملاحظات مديريه",
        "9- رفع التقارير وإخبار مديره بالمعلومات ونقل المشاكل والفيدباك بشكل مستمر",
        "10- الالتزام بمهام التشطيب وجودة أداء المهام المطلوبة",
        "11- التباهي بالسلوكيات الخاطئة (سهرات او انشطة مشبوهة بعد الدوام)",
        "12- مدى التركيز والانتباه على الصالة وراحة العملاء وأي مشاكل أو خلل وسرعة التعامل معها وإيصالها و التركيز على حلها وعدم نسيانها",
        "13- مدى الالتزام بالألقاب مع المسؤول (سير)",
        "14- سرعة أخذ الطلبات وإنجازها 2-3 دقائق والتأكد من نظافة الطلب ومطابقة المعايير من حيث نظافة الكاسة والالتزام بالريسيبي",
        "15- تشطيب البار وتنظيفه خلال 45 دقيقة",
        "16- استقبال الضيوف وتشمل الموظفين القدامى",
        "17- حفظ أسماء العملاء وألقابهم مع الحفاظ على طبيعة العلاقة friendly ليس friends والالتزام والحفاظ على العلاقات الرسمية معهم ومدة الوقوف معهم وتجنب الأحاديث الجانبية",
        "18- الحفاظ على الوقفة السليمة والترحيب بالعملاء",
        "19- مراقبة الطلبات المرتجعة أو لم يتم شربها وعمل تقرير بها ومعرفة المشكلة",
        "20- ضبط جودة المشروبات والطلبات عند التحضير وقبل التسليم",
        "21- مدى الاتكالية في المهام وسوء استخدام السلطة مع الموظفين",
        "22- فهم الواجبات والمهام المطلوبة منه بشكل واضح وكيفية تطبيقها و القدرة على تدريب موظفيه ورفع أدائهم ومعرفتهم في الباريستا",
        "23- الصيانات داخل البار ومتابعتها حتى يتم إنجازها",
        "24- القدرة على طلب الطلبيات والحفاظ على الـstock المناسب ومتابعة المخزون وقت وصول الطلبية والإبلاغ في حال كان هناك نقص أو مشكلة أو تأخير في وقت الوصول",
        "25- قدرة الموظفين على العمل دون إشراف ورقابة وتوجيه دائم ضبط سلوكهم وتحركاتهم داخل البار",
        "26- القدرة على حل المشاكل سواء مشاكل الموظفين الخاصة أو ومشاكل بين الموظفين خلال العمل وتوزيع البريكات على الموظفين بشكل عادل"
    ]


    def get_columns_for_role(role):
        return [col for col in df.columns if role in col]

    additional_columns = ['Branch', 'اسم المسؤول السوبرفايزر', 'الشفت', 'Submission started']

    # Ensure the additional columns exist in the DataFrame
    additional_columns = [col for col in additional_columns if col in df.columns]

    def create_role_data(index, max_length, role_columns):
        role_data = [None] * max_length
        for col in role_columns:
            try:
                col_index = int(col.split('-')[0].strip())
                if col_index <= max_length:
                    role_data[col_index - 1] = df.at[index, col]
            except ValueError:
                continue
        return role_data

    def sanitize_sheet_name(name):
        return str(name).replace(':', '-').replace('/', '-').replace('\\', '-').replace('*', '-').replace('?',
                                                                                                          '-').replace(
            '[', '-').replace(']', '-')

    output_path_custom = 'kpi.xlsx'

    with pd.ExcelWriter(output_path_custom, engine='openpyxl') as writer:
        sheet_created = False
        for index, row in df.iterrows():
            try:
                sheet_name = sanitize_sheet_name(row['اسم المسؤول السوبرفايزر'])

                max_length = len(kpi_criteria)
                data = {'KPI Criteria': kpi_criteria}

                for role in roles:
                    role_columns = get_columns_for_role(role)
                    role_data = create_role_data(index, max_length, role_columns)
                    data[role] = role_data

                restructured_df = pd.DataFrame(data)

                averages = (restructured_df.apply(pd.to_numeric, errors='coerce').mean()) / 0.03
                averages_row = pd.DataFrame(averages).T

                if additional_columns:
                    additional_data = row[additional_columns].values
                    additional_df = pd.DataFrame([additional_data], columns=additional_columns)
                    additional_df.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)

                dynamic_header = [''] + [row['اسم الهيد'], row['اسم الكاشير'], row['اسم باريستا(1)'], row['اسم باريستا(2)'],
                                             row['اسم موظف صالة(1)'], row['اسم باريستا(3)'], row['اسم موظف صالة(2)'],
                                             row['اسم موظف صالة(3)'], row['اسم باريستا(4)']]
                header_df = pd.DataFrame([dynamic_header])
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False, header=False)

                restructured_df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)

                averages_row.to_excel(writer, sheet_name=sheet_name, startrow=len(restructured_df) + 3, header=False,
                                      index=False)

                worksheet = writer.sheets[sheet_name]

                # Set modern font and alignment
                modern_font = Font(name='Calibri', size=12, bold=True)
                header_font = Font(name='Calibri', size=12, bold=True)
                alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))

                # Color header
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                for col_index, cell in enumerate(worksheet[2], start=1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border

                # Color additional columns
                for col_index, col_name in enumerate(additional_columns, start=1):
                    cell = worksheet.cell(row=1, column=col_index)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border
                    value = row[col_name]
                    value_cell = worksheet.cell(row=2, column=col_index)
                    value_cell.value = value
                    value_cell.font = modern_font
                    value_cell.alignment = alignment
                    value_cell.border = border

                # Format all other cells
                for row in worksheet.iter_rows(min_row=3, max_row=len(restructured_df) + 3, min_col=1,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.font = modern_font
                        cell.alignment = alignment
                        cell.border = border

                # Align Index column to the left
                for cell in worksheet['A']:
                    cell.alignment = left_alignment

                highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for cell in worksheet.iter_rows(min_row=len(restructured_df) + 4, max_row=len(restructured_df) + 4,
                                                min_col=1, max_col=worksheet.max_column):
                    for c in cell:
                        c.fill = highlight_fill
                        c.font = modern_font
                        c.alignment = alignment
                        c.border = border

                # Autofit columns
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

                sheet_created = True
            except Exception as e:
                print(f"An error occurred while processing row {index}: {e}")
                continue

        if not sheet_created:
            # Create an empty sheet to avoid IndexError
            pd.DataFrame().to_excel(writer, sheet_name='EmptySheet')

    return output_path_custom





@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(file_path)
            try:
                restructured_file_path = restructure_excel(file_path)
                return send_file(restructured_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return '''
<!doctype html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>رفع ملف جديد</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f5f5f5;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin: 0;
        }
        .container {
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            text-align: center;
        }
        h1, h2 {
            margin: 0 0 20px 0;
        }
        h2 strong {
            color: red;
        }
        form {
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        input[type="file"] {
            background-color: tomato;
            color: white;
            border: none;
            padding:5px 20px 5px 20px;
            margin:50px;
            width:400px;
            font-size: 20px;
            border-radius: 4px;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
         input[type="file"]:hover {
            background-color: black;

        }
        button[type="submit"] {
            background-color: #007bff;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            height:50px;
            padding:5px 190px 5px 190px;
            font-size:14px;
            transition: background-color 0.3s ease;
        }
        button[type="submit"]:hover {
            background-color: #0056b3;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>رفع ملف جديد</h1>
        <h2>تأكد من أن صيغة الجدول هي <strong>.xlsx</strong> قبل الرفع.</h2>
        <form method="post" enctype="multipart/form-data">
           <p>ارفق الملف لتحويله</p> <input type="file" name="file" accept=".xlsx" >
            <button type="submit">تحميل</button>
        </form>
    </div>
</body>
</html>

    '''


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5000,)
from flask import Flask, request, send_file, render_template_string
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def restructure_excel(file_path):
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls)

    # Define roles
    roles = ['هيد', 'كاشير', 'باريستا(1)', 'باريستا(2)', 'باريستا(3)', 'باريستا(4)', 'صالة(1)', 'صالة(2)', 'صالة(3)']

    kpi_criteria = [
        "1- الالتزام بمواعيد الحضور والغياب",
        "2- الالتزام باليونيفورم واللباس المطلوب من حيث نظافته ، العطر , الأظافر، الشعر ، النظافة الشخصية ، والاكسسوارات",
        "3- الالتزام بمواعيد البريكات وترك غرفة البريك نظيفة",
        "4- عدم استخدام الهاتف الشخصي وان يكون هاتفه متاح بعد الدوام",
        "5- هل يحترم زملائه ويجيد التعامل معهم ولا يتسبب في حدوث مشاكل بين زملائه",
        "6- هل يحتاج إلى التوجيه والرقابة المستمرة أو قادر على إنجاز أعماله دون انتظار التعليمات والتوجيهات",
        "7- هل تطورت مهاراته الوظيفية من حيث السرعة والإتقان",
        "8- هل يتأقلم ويتقبل التعديلات والتغييرات بسرعة؟ إلى أي مدى يتقبل ويستوعب تعليقات وملاحظات مديريه",
        "9- رفع التقارير وإخبار مديره بالمعلومات ونقل المشاكل والفيدباك بشكل مستمر",
        "10- الالتزام بمهام التشطيب وجودة أداء المهام المطلوبة",
        "11- التباهي بالسلوكيات الخاطئة (سهرات او انشطة مشبوهة بعد الدوام)",
        "12- مدى التركيز والانتباه على الصالة وراحة العملاء وأي مشاكل أو خلل وسرعة التعامل معها وإيصالها و التركيز على حلها وعدم نسيانها",
        "13- مدى الالتزام بالألقاب مع المسؤول (سير)",
        "14- سرعة أخذ الطلبات وإنجازها 2-3 دقائق والتأكد من نظافة الطلب ومطابقة المعايير من حيث نظافة الكاسة والالتزام بالريسيبي",
        "15- تشطيب البار وتنظيفه خلال 45 دقيقة",
        "16- استقبال الضيوف وتشمل الموظفين القدامى",
        "17- حفظ أسماء العملاء وألقابهم مع الحفاظ على طبيعة العلاقة friendly ليس friends والالتزام والحفاظ على العلاقات الرسمية معهم ومدة الوقوف معهم وتجنب الأحاديث الجانبية",
        "18- الحفاظ على الوقفة السليمة والترحيب بالعملاء",
        "19- مراقبة الطلبات المرتجعة أو لم يتم شربها وعمل تقرير بها ومعرفة المشكلة",
        "20- ضبط جودة المشروبات والطلبات عند التحضير وقبل التسليم",
        "21- مدى الاتكالية في المهام وسوء استخدام السلطة مع الموظفين",
        "22- فهم الواجبات والمهام المطلوبة منه بشكل واضح وكيفية تطبيقها و القدرة على تدريب موظفيه ورفع أدائهم ومعرفتهم في الباريستا",
        "23- الصيانات داخل البار ومتابعتها حتى يتم إنجازها",
        "24- القدرة على طلب الطلبيات والحفاظ على الـstock المناسب ومتابعة المخزون وقت وصول الطلبية والإبلاغ في حال كان هناك نقص أو مشكلة أو تأخير في وقت الوصول",
        "25- قدرة الموظفين على العمل دون إشراف ورقابة وتوجيه دائم ضبط سلوكهم وتحركاتهم داخل البار",
        "26- القدرة على حل المشاكل سواء مشاكل الموظفين الخاصة أو ومشاكل بين الموظفين خلال العمل وتوزيع البريكات على الموظفين بشكل عادل"
    ]


    def get_columns_for_role(role):
        return [col for col in df.columns if role in col]

    additional_columns = ['Branch', 'اسم المسؤول السوبرفايزر', 'الشفت', 'Submission started']

    # Ensure the additional columns exist in the DataFrame
    additional_columns = [col for col in additional_columns if col in df.columns]

    def create_role_data(index, max_length, role_columns):
        role_data = [None] * max_length
        for col in role_columns:
            try:
                col_index = int(col.split('-')[0].strip())
                if col_index <= max_length:
                    role_data[col_index - 1] = df.at[index, col]
            except ValueError:
                continue
        return role_data

    def sanitize_sheet_name(name):
        return str(name).replace(':', '-').replace('/', '-').replace('\\', '-').replace('*', '-').replace('?',
                                                                                                          '-').replace(
            '[', '-').replace(']', '-')

    output_path_custom = 'kpi.xlsx'

    with pd.ExcelWriter(output_path_custom, engine='openpyxl') as writer:
        sheet_created = False
        for index, row in df.iterrows():
            try:
                sheet_name = sanitize_sheet_name(row['اسم المسؤول السوبرفايزر'])

                max_length = len(kpi_criteria)
                data = {'KPI Criteria': kpi_criteria}

                for role in roles:
                    role_columns = get_columns_for_role(role)
                    role_data = create_role_data(index, max_length, role_columns)
                    data[role] = role_data

                restructured_df = pd.DataFrame(data)

                averages = (restructured_df.apply(pd.to_numeric, errors='coerce').mean()) / 0.03
                averages_row = pd.DataFrame(averages).T

                if additional_columns:
                    additional_data = row[additional_columns].values
                    additional_df = pd.DataFrame([additional_data], columns=additional_columns)
                    additional_df.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)

                dynamic_header = [''] + [row['اسم الهيد'], row['اسم الكاشير'], row['اسم باريستا(1)'], row['اسم باريستا(2)'],
                                             row['اسم موظف صالة(1)'], row['اسم باريستا(3)'], row['اسم موظف صالة(2)'],
                                             row['اسم موظف صالة(3)'], row['اسم باريستا(4)']]
                header_df = pd.DataFrame([dynamic_header])
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False, header=False)

                restructured_df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)

                averages_row.to_excel(writer, sheet_name=sheet_name, startrow=len(restructured_df) + 3, header=False,
                                      index=False)

                worksheet = writer.sheets[sheet_name]

                # Set modern font and alignment
                modern_font = Font(name='Calibri', size=12, bold=True)
                header_font = Font(name='Calibri', size=12, bold=True)
                alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))

                # Color header
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                for col_index, cell in enumerate(worksheet[2], start=1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border

                # Color additional columns
                for col_index, col_name in enumerate(additional_columns, start=1):
                    cell = worksheet.cell(row=1, column=col_index)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border
                    value = row[col_name]
                    value_cell = worksheet.cell(row=2, column=col_index)
                    value_cell.value = value
                    value_cell.font = modern_font
                    value_cell.alignment = alignment
                    value_cell.border = border

                # Format all other cells
                for row in worksheet.iter_rows(min_row=3, max_row=len(restructured_df) + 3, min_col=1,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.font = modern_font
                        cell.alignment = alignment
                        cell.border = border

                # Align Index column to the left
                for cell in worksheet['A']:
                    cell.alignment = left_alignment

                highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for cell in worksheet.iter_rows(min_row=len(restructured_df) + 4, max_row=len(restructured_df) + 4,
                                                min_col=1, max_col=worksheet.max_column):
                    for c in cell:
                        c.fill = highlight_fill
                        c.font = modern_font
                        c.alignment = alignment
                        c.border = border

                # Autofit columns
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

                sheet_created = True
            except Exception as e:
                print(f"An error occurred while processing row {index}: {e}")
                continue

        if not sheet_created:
            # Create an empty sheet to avoid IndexError
            pd.DataFrame().to_excel(writer, sheet_name='EmptySheet')

    return output_path_custom





@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(file_path)
            try:
                restructured_file_path = restructure_excel(file_path)
                return send_file(restructured_file_path, as_attachment=True)
            except Exception as e:
                return f"An error occurred: {e}"
        else:
            return 'Invalid file type. Only .xlsx files are allowed.'
    return '''
<!doctype html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>رفع ملف جديد</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f5f5f5;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin: 0;
        }
        .container {
            background-color: #fff;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            text-align: center;
        }
        h1, h2 {
            margin: 0 0 20px 0;
        }
        h2 strong {
            color: red;
        }
        form {
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        input[type="file"] {
            background-color: tomato;
            color: white;
            border: none;
            padding:5px 20px 5px 20px;
            margin:50px;
            width:400px;
            font-size: 20px;
            border-radius: 4px;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
         input[type="file"]:hover {
            background-color: black;

        }
        button[type="submit"] {
            background-color: #007bff;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            height:50px;
            padding:5px 190px 5px 190px;
            font-size:14px;
            transition: background-color 0.3s ease;
        }
        button[type="submit"]:hover {
            background-color: #0056b3;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>رفع ملف جديد</h1>
        <h2>تأكد من أن صيغة الجدول هي <strong>.xlsx</strong> قبل الرفع.</h2>
        <form method="post" enctype="multipart/form-data">
           <p>ارفق الملف لتحويله</p> <input type="file" name="file" accept=".xlsx" >
            <button type="submit">تحميل</button>
        </form>
    </div>
</body>
</html>

    '''


if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5000,)
